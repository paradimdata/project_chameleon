import os.path
import os
import re
import time
import htmdec_formats
import h5py
import subprocess
import traceback
import argparse
from htmdec_formats import ARPESDataset
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from configparser import ParsingError
from pandas import DataFrame
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import QApplication, QLabel, QTextEdit, QVBoxLayout, QWidget

def get_sec(time_str):
    """Get seconds from time."""
    h, m, s = time_str.split(':')
    return int(h) * 3600 + int(m) * 60 + int(s)


def build_arpes_workbook(workbook_name):
    """
    ``build_arpes_workbook`` is a function that creates an excel .xlsx file with a specific premade format. This format is used to hold ARPES data. The function is used to create a new workbook for data to be put into in the ``arpes_folder_workbook`` function.

    :args: ``workbook_name`` is a string or path that must end with '.xlsx'.

    :return: Does not return anything. Creates an excel file with a premade format. 

    :exceptions: ``workbook_name`` must end with '.xslx'.
    
    """

    # File must be an excel file so openpyxl can be used
    if not str(workbook_name).endswith('.xlsx'):
        raise ValueError("ERROR: workbook_name input must end with '.xlsx'")
    
    # Open the workbook so we have an open workbook to work in
    wb = Workbook()
    ws = wb.active

    # Row labels are set so they can be input later
    row_1 = ['Colored Labels','','','Colored Labels','Alignment Maps','Hi-Stat Map','"Good" High-Stat Scan','"Good" High-Stat Scan','Mono Lamp Change','Mono Lamp Change',
             'Slit Change','Slit Change','Theta Offset:','0','Phi Offset:','0','Omega Offset:','0']
    row_2 = ['Scan Number/File Name','Date','Time Start', 'Time End', 'Notes/Comments','FS Position','(X,Y,Z)','Theta(encoder)','Theta(true)','Phi(encoder)','Phi(true)',
             'Omega(Manipulator,approx)','Omega(true)','Kinetic Energy Range(eV)','Step Size(eV)','Temperature(K)[Diode A,Diode B]','Run Mode','Acquisition Mode',
             '[# of Sweeps, Layers]','Pass Energy','Analyzer Slit','Pressure(Torr)','Photon Energy(eV)', 'Full Notes/Comments']
    
    # Column values are set so we know how long to set for loops to
    first_col = 1
    last_col = 28
    row_1_last_col = 19
    row_2_last_col = 25
    end_col = 14

    # Border is style is set to thin because that fits our design
    thin_border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin')
    ) 

    # Set column widths are set wider so more information can fit inside them
    for col in range(first_col, last_col + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

    # Set borders on cells so division on cells can be easily seen
    for col in range(first_col, end_col + 1):
            cell = ws.cell(row=2, column=col)
            cell.border = thin_border
    
    # Put values from row_1 into cells. Align them vertically and make them bold to make them more easily readable
    for col in range(first_col, row_1_last_col):
        if row_1[col - 1] == '':
            continue
        else:
            cell = ws.cell(row=1, column=col)
            cell.value = row_1[col-1] 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Put values from row_2 into cells. Align them vertically and make them bold to make them more easily readable
    for col in range(first_col, row_2_last_col):
        cell = ws.cell(row=2, column=col)
        cell.value = row_2[col-1] 
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set background colors of cells so cells can be easily distiguished
    fill_color1 = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")#Light orange
    fill_color2 = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")#Light purple
    fill_color3 = PatternFill(start_color="66CDAA", end_color="66CDAA", fill_type="solid")#Light green
    fill_color4 = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")#Light blue
    fill_color5 = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")#light yellow
    fill_color6 = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")#light red
    fill_grey = fill_color = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Merge cells so labels cover correct cells
    ws.merge_cells('A1:B1')
    ws.merge_cells('G1:H1')
    ws.merge_cells('I1:J1')
    ws.merge_cells('K1:L1')
    ws.merge_cells('A3:A4')
    ws.merge_cells('B3:I4')

    # Apply colors to cells so cells can be more easily distiguished
    ws['E1'].fill = fill_color1
    ws['F1'].fill = fill_color2
    ws['G1'].fill = fill_color3
    ws['I1'].fill = fill_color4
    ws['K1'].fill = fill_color5
    ws['A5'].fill = fill_color6
    ws['M1'].fill = fill_grey

    # Format additional cells that cannot be formated within for loops. 
    cell = ws['A3']
    cell.value = 'Initial Notes/Sample Information'
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell = ws['A5']
    cell.value = 'Initial Start:'
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save and close workbook so changes are saved and workbook is not left open
    wb.save(workbook_name)


def get_wavenote_values(wavenote_file):
    """
    ``get_wavenote_values`` is a function that extracts values from the wave note of the .pxt file using HTMDEC_formats. This function does not extract all values, only values that were consiedered relevant. All relevant values are returned. Which values are considered relevant can be easily changed locally. 

    :args: ``wavenote_file`` is a string or path to a '.pxt' file.

    :return: Returns a 1 dimensional array that holds relevant values from the wave note of the .pxt file.

    :exceptions: ``wavenote_file`` must end with '.pxt'. ``wavenote_file`` must be a file.
    
    """
    # Validate input file
    if not wavenote_file.endswith('.pxt'):
        raise ValueError("ERROR: Input file must end with '.pxt'")
    if not os.path.isfile(wavenote_file):
        raise ValueError("ERROR: File not found")

    try:
        # Get file metadata
        ti_m = os.path.getmtime(wavenote_file)
        end_time = time.strftime("%H:%M:%S", time.localtime(ti_m))  # Format as HH:MM:SS
        
        # Load dataset and split metadata into lines
        dataset = ARPESDataset.from_file(wavenote_file)
        lines = dataset._metadata.split("\n")
        
        # Extract scan number from lines[20]
        scan_number = next((l for l in lines[20].split('\\') if '.pxt' in l), None)
        
        # Determine scan type based on specific characteristics
        if lines[35]:
            scan_type = 'Add Dimension'
        elif len(dataset.run_mode_info) > 0:
            scan_type = 'Manipulator Scan'
        else:
            scan_type = 'Normal Mode'

        # Extract and format relevant data
        result = [
            scan_number,
            lines[28].split('=', 1)[-1].strip(),  # Start Date
            lines[29].split('=', 1)[-1].strip(),  # Start Time
            end_time,  # End Time
            lines[27].split('=', 1)[-1].strip(),  # Comments
            f"[{lines[11].split('=', 1)[-1].strip()}, {lines[12].split('=', 1)[-1].strip()}]",  # Kinetic Energy Range
            lines[13].split('=', 1)[-1].strip(),  # Step Size
            scan_type,
            lines[8].split('=', 1)[-1].strip(),  # Acquisition Mode
            f"[{lines[5].split('=', 1)[-1].strip()}, {dataset.num_layers}]",  # # of sweeps
            lines[4].split('=', 1)[-1].strip(),  # Pass Energy
            lines[6].split('=', 1)[-1].strip(),  # Photon Energy
        ]
        
        return result

    except ParsingError as e:
        print(f"Parsing error in get_wavenote_values: {e}")
        traceback.print_exc()
    except Exception as e:
        print(f"Unexpected error in get_wavenote_values: {e}")
        traceback.print_exc()



def get_varian_values(varian_file, date_time = None):
    """
    ``get_varian_values`` is a function that extracts values from the 'varian' .log file. This function does not extract all values, only relevant values. This function was designed for the 'varian' logs generated from the ARPES machine on which this function was designed. 

    :args: ``varian_file`` is a string or path that end with '.log'.``date_time`` is an array of length 2, that only holds a date and a time. ``date_time`` is an optional input parameter.

    :return: Returns a 1 dimensional array that holds relevant values from the varian file.

    :exceptions: ``varian_file`` must end with '.log'. ``varian_file`` must be a file.
    
    """
    # Formatted for extracting data from a .log file so file needs to be a .log file
    if not varian_file.endswith('.log'):
        raise ValueError("ERROR: input file must end with '.log'")
    if os.path.isfile(varian_file) is False:
        raise ValueError("ERROR: bad input. Expected file")
    
    # Variables set here so they can be added to later
    varian_lines = []
    time_found = 0

    # Make sure the date in date_time matches the date of the log file. If not, find the log file that does match so we get the value from when the scan actually happened
    if date_time:
        varian_folder = os.listdir(os.path.dirname(varian_file))
        # Make dates the same format
        if '/' in varian_file:
            varian_name = varian_file.split('/')[-1].split('.')[0]
        else: 
            varian_name = varian_file.split('.')[0]
        # Date in dd-mm-yyyy format
        varian_name = varian_name.replace('-','/')
        if varian_name != date_time[0]:
            # Go through each log file in the folder to check if they have the correct date. Each file only has one date
            for item in varian_folder:
                if '.log' in item:
                    if item.split('.')[0].replace('-','/') == date_time[0]:
                        varian_file = os.path.join(os.path.dirname(varian_file), item)
                        break
                    else:
                        continue
                        
    # Open log, read lines after 1, 1 is just titles of coluns
    with open(varian_file, 'r') as file:
        varian_lines = file.readlines()
    varian_lines = varian_lines[1:]

    # No date time given, just take first row because we have no date to file
    if date_time == None:
        # [(x,y,z),Theta,Phi,Analyzer Slit]
        return ['(' + str(Decimal(varian_lines[1].replace(' ','').replace('\n','').split("\t")[2]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ',' + 
                        str(Decimal(varian_lines[1].replace(' ','').replace('\n','').split("\t")[4]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ',' + 
                        str(Decimal(varian_lines[1].replace(' ','').replace('\n','').split("\t")[6]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ')',
                        varian_lines[1].replace(' ','').replace('\n','').split("\t")[8],varian_lines[1].replace(' ','').replace('\n','').split("\t")[10],
                        varian_lines[1].replace(' ','').replace('\n','').split("\t")[14]]
    
    # If there is a date time, find which line in the file has the closest time to the date_time so we can get the most accurate value
    else:
        while time_found < 1:
            length = len(varian_lines)
            # Once length is 5 or less just look through all the items. Should avoid missing on edge cases
            if length <= 5:
                if length == 0:
                    # Break code if we can't find a file with the correct time because something wrong
                    raise ValueError("ERROR: Error in varian log file. Time cannot be found")
                for item in varian_lines:
                    if (get_sec(date_time[1]) >= get_sec(item.replace(' ','').replace('\n','').split("\t")[0][10:18]) - 30) and (get_sec(date_time[1]) <= get_sec(item.replace(' ','').replace('\n','').split("\t")[0][10:18]) + 30):
                        return ['(' + str(Decimal(item.replace(' ','').replace('\n','').split("\t")[2]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ',' + 
                        str(Decimal(item.replace(' ','').replace('\n','').split("\t")[4]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ',' + 
                        str(Decimal(item.replace(' ','').replace('\n','').split("\t")[6]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ')',
                        item.replace(' ','').replace('\n','').split("\t")[8],item.replace(' ','').replace('\n','').split("\t")[10],
                        item.replace(' ','').replace('\n','').split("\t")[14]]
            middle = length // 2

            # If the time is more than 30 seconds greater than the middle, take higher half of array
            if (get_sec(date_time[1]) > get_sec(varian_lines[middle].replace(' ','').replace('\n','').split("\t")[0][10:18]) + 30): 
                varian_lines = varian_lines[middle:]

            # If the time is more than 30 seconds less than the middle, take lower half of array
            elif (get_sec(date_time[1]) < get_sec(varian_lines[middle].replace(' ','').replace('\n','').split("\t")[0][10:18]) - 30):
                varian_lines = varian_lines[:middle]

            # If the time is within 30 seconds on either side, extract the values, then return the values
            elif (get_sec(date_time[1]) >= get_sec(varian_lines[middle].replace(' ','').replace('\n','').split("\t")[0][10:18]) - 30) and (get_sec(date_time[1]) <= get_sec(varian_lines[middle].replace(' ','').replace('\n','').split("\t")[0][10:18]) + 30):
                return ['(' + str(Decimal(varian_lines[middle].replace(' ','').replace('\n','').split("\t")[2]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ',' + 
                        str(Decimal(varian_lines[middle].replace(' ','').replace('\n','').split("\t")[4]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ',' + 
                        str(Decimal(varian_lines[middle].replace(' ','').replace('\n','').split("\t")[6]).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)) + ')',
                        varian_lines[middle].replace(' ','').replace('\n','').split("\t")[8],varian_lines[middle].replace(' ','').replace('\n','').split("\t")[10],
                        varian_lines[middle].replace(' ','').replace('\n','').split("\t")[14]]


def get_jaina_values(jaina_file, date_time = None):
    """
    ``get_jaina_values`` is a function that extracts values from the 'jaina' .log file. This function does not extract all values, only relevant values. This function was designed for the 'jaina' logs generated from the ARPES machine on which this function was designed. 

    :args: ``jaina_file`` is a string or path that end with '.log'.``date_time`` is an array of length 2, that holds a date and a time. ``date_time`` is an optional input parameter.

    :return: Returns a 1 dimensional array that holds relevant values from the varian file.

    :exceptions: ``jaina_file`` must end with '.log'. ``jaina_file`` must be a file.
    
    """

    # Formatted for extracting data from a .log file so file needs to be a .log file
    if not jaina_file.endswith('.log'):
        raise ValueError("ERROR: input file must end with '.log'")
    if os.path.isfile(jaina_file) is False:
        raise ValueError("ERROR: bad input. Expected file")
    
    # Set values here so they can added to later
    jaina_lines = []
    time_found = 0
    
    # Make sure the date in date_time matches the date of the log file. If not, find the log file that does match so we get the value from when the scan actually happened
    if date_time:
        jaina_folder = os.listdir(os.path.dirname(jaina_file))
        # Make dates the same format
        if '/' in jaina_file:
            jaina_name = jaina_file.split('/')[-1].split('.')[0]
        else: 
            jaina_name = jaina_file.split('.')[0]
        # Date in dd-mm-yyyy format
        jaina_name = jaina_name.replace('-','/')
        if jaina_name != date_time[0]:
            # Go through each log file in the folder to check if they have the correct date. Each file only has one date
            for item in jaina_folder:
                if '.log' in item:
                    if item.split('.')[0].replace('-','/') == date_time[0]:
                        jaina_file = os.path.join(os.path.dirname(jaina_file), item)
                        break
                    else:
                        continue

    # Open log, read lines after 1, 1 is just titles of coluns
    with open(jaina_file, 'r') as file:
        jaina_lines = file.readlines()
    jaina_lines = jaina_lines[1:]

    # If time is given, align varian values to time, otherwise take first row
    if date_time == None:
        # [(Diode A,Diode B), Pressure]
        return ['[' + jaina_lines[1].replace(' ','').replace('\n','').split("\t")[1] + ',' + jaina_lines[1].replace(' ','').replace('\n','').split("\t")[2] + ']',jaina_lines[1].replace(' ','').replace('\n','').split("\t")[14]]
    
    # If there is a date time, find which line in the file has the closest time to the date_time so we can get the most accurate value
    else:
        # Binary search method to find the time closest to the time submitted in date_time
        while time_found < 1:
            length = len(jaina_lines)
            # Once length is 5 or less just look through all the items. Should avoid missing on edge cases
            if length <= 5:
                if length == 0:
                    # Break code if we can't find a file with the correct time because something wrong
                    raise ValueError("ERROR: Error in jaina log file. Time cannot be found")
                for item in jaina_lines:
                    if (get_sec(date_time[1]) >= get_sec(item.replace(' ','').replace('\n','').split("\t")[0][10:18]) - 30) and (get_sec(date_time[1]) <= get_sec(item.replace(' ','').replace('\n','').split("\t")[0][10:18]) + 30):
                        return ['[' + item.replace(' ','').replace('\n','').split("\t")[1] + ',' + 
                                item.replace(' ','').replace('\n','').split("\t")[2] + ']',item.replace(' ','').replace('\n','').split("\t")[14]]
            middle = length // 2

            # If the time is more than 30 seconds greater than the middle, take higher half of array
            if (get_sec(date_time[1]) > get_sec(jaina_lines[middle].replace(' ','').replace('\n','').split("\t")[0][10:18]) + 30): 
                jaina_lines = jaina_lines[middle:]

            # If the time is more than 30 seconds less than the middle, take lower half of array
            elif (get_sec(date_time[1]) < get_sec(jaina_lines[middle].replace(' ','').replace('\n','').split("\t")[0][10:18]) - 30):
                jaina_lines = jaina_lines[:middle]

            # If the time is within 30 seconds on either side, extract the values, then return the values
            elif (get_sec(date_time[1]) >= get_sec(jaina_lines[middle].replace(' ','').replace('\n','').split("\t")[0][10:18]) - 30) and (get_sec(date_time[1]) <= get_sec(jaina_lines[middle].replace(' ','').replace('\n','').split("\t")[0][10:18]) + 30):
                return ['[' + jaina_lines[middle].replace(' ','').replace('\n','').split("\t")[1] + ',' + jaina_lines[middle].replace(' ','').replace('\n','').split("\t")[2] + ']',jaina_lines[middle].replace(' ','').replace('\n','').split("\t")[14]]


def insert_scan_row(wavenote_file,jaina_file,varian_file,workbook_name):
    """
    ``insert_scan_row`` is a function that extracts relevant values from a .pxt wave note file, a jaina log file, a varian log file, and inserts all of the values into the first open row of the workbook ``workbook_name``. This function utilizes the functionality of ``get_wavenote_values``, ``get_jaina_values``, and ``get_varian_values``.

    :args: ``wavenote_file`` is a string or a path. ``jaina_file`` is a string or path. ``varian_file`` is a string or a path. ``workbook_name`` is a string or a path.

    :return: Does not return anything. Inserts values into the workbook ``workbook_name``.

    :exceptions: ``wavenote_file`` must end with '.pxt'. ``jaina_file`` must end with '.log'. ``varian_file`` must end with '.log'. ``workbook_name`` must end with '.xlsx'.
    
    """
    # Make sure all the input files are of the types expected
    if not str(workbook_name).endswith('.xlsx'):
        raise ValueError("ERROR: workbook_name input must end with '.xlsx'")
    if not jaina_file.endswith('.log'):
        raise ValueError("ERROR: jaina_file must end with '.log'")
    if not varian_file.endswith('.log'):
        raise ValueError("ERROR: varian_file must end with '.log'")
    if not wavenote_file.endswith('.pxt'):
        raise ValueError("ERROR: wavenote_file must end with '.pxt'")

    # Set columns so for loops are the correct length
    first_col = 1
    last_col = 25

    # Set staring row based on known excel format, set starting cell to 0 so we know the starting cell hasnt been found
    starting_row = 6
    starting_cell = 0

    delimiters = r"[;,]"
    old_comment = ''
    old_split = []
    final_comment = []

    light_orange_fill = PatternFill(start_color="FFE5B2", end_color="FFE5B2", fill_type="solid")

    # Read wavenote, extract start and end time for jaina,varian if wavenote does not return error
    w = get_wavenote_values(wavenote_file)
    if w != None:
        date = w[1]
        time = w[2]
        date = date.replace('-','/').replace('\n','').replace('\n','')
        time = time.replace('\n','')
        j = get_jaina_values(jaina_file,[date,time])
        v = get_varian_values(varian_file,[date,time])

    # Make the workbook and open it so we have somewhere to put the data
    wb = wb = load_workbook(filename = workbook_name)
    ws = wb.active

    # Find first open row by checking if the cell has a value. If there is no cell value loop breaks and we know where our starting row is
    while starting_cell == 0:
        cell = ws.cell(row=starting_row, column=first_col)
        if not cell.value:
            starting_cell = 1
        else:
            starting_row += 1

    # Check comments. Only keep new information in comments
    if starting_row > 6:
        old_comment = ws['E6'].value
        old_split = re.split(delimiters, old_comment)
    new_comment = w[4]
    new_split = re.split(delimiters, new_comment)

    for e in new_split:
        if not e in old_split:
            final_comment.append(e)

    final_comment = ', '.join(final_comment)
    if final_comment == '':
        final_comment = new_comment



    # Make the each cell in the row read "Error" if wavenote cant be read
    if w == None:
        ws.row_dimensions[starting_row].height = 40
        for col in range(first_col, last_col):
            cell = ws.cell(row=starting_row, column=1)
            cell.value = 'Error' 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

    else:
        # Insert all data into a 1D array in the format that fits predefined excel sheet format
        # [Scan, Date, Start time, End time, Notes/Comments, '', (X,Y,Z), Theta, '', Phi, '', '', '', Kinetic Energy, Step size, Temp[A,B], Run mode, Acquisition mode, # # of sweeps, Pass energy, Analyzer slit, Pressure, Photon Energy]
        if w[7] == 'Manipulator Scan':
            dataset = htmdec_formats.ARPESDataset.from_file(wavenote_file)
            df = pd.DataFrame(dataset.run_mode_info)
            v[1] = str((float(df.iloc[0,4]),float(df.iloc[-1,4])))
            v[2] = str((float(df.iloc[0,5]),float(df.iloc[-1,5])))

        row = [w[0],w[1],w[2],w[3],final_comment,'',v[0],v[1],'',v[2],'','','',w[5],w[6],j[0],w[7],w[8],w[9],w[10],v[3],j[1],w[11],w[4]]

        # Insert values into rows, set height of cells larger so more data can fit in them
        ws.row_dimensions[starting_row].height = 40
        for col in range(first_col, last_col):

            # For the notes column make the font smaller so all of the notes can fit into the cell
            if col == 5:
                cell = ws.cell(row=starting_row, column=col)
                cell.value = row[col-1] 
                cell.font = Font(bold=False, size = 8)
                cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

            elif col == 24:
                cell = ws.cell(row=starting_row, column=col)
                cell.value = row[col-1] 
                cell.font = Font(bold=False, size = 8)
                cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

            # Otherwise insert data normally
            else:
                cell = ws.cell(row=starting_row, column=col)
                cell.value = row[col-1] 
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center',wrapText=True)

        if w[7] == 'Manipulator Scan':
            for i in range(last_col):
                ws[f'{chr(65 + i)}{starting_row}'].fill = light_orange_fill
                cell = ws.cell(row=starting_row, column=i + 1)
                cell.border = Border(top=Side(style='thin'),bottom=Side(style='thin'),left=Side(style='thin'),right=Side(style='thin')) 

    # Close and save workbook so all data added is saved and workbook isn't altered further
    wb.save(workbook_name)

    return v[3]


def arpes_folder_workbook(folder_name, workbook_name):
    """
    ``arpes_folder_workbook`` is a function that creates a workbook, looks through the folder to see how many scans there are, extracts values from .pxt, jaina, and varian files for each scan, and creates a row with the extracted values for each scan. This function utilizes functionality of  ``build_arpes_workbook`` and ``insert_scan_row``.

    :args: ``folder_name`` is a string or a path. ``workbook_name`` is a string or a path that must end with '.xlsx'.

    :return: Does not return anything. Creates a workbook with a row for each .pxt file in the folder.

    :exceptions: ``folder_name`` must be a directory. ``workbook_name`` must end with '.xlsx'.
    
    """

    # Make sure the folder input is actually a folder with files so the function can actually run. Also make sure the workbook is acutally an excel file 
    if not os.path.isdir(folder_name):
        raise ValueError("ERROR: folder_name input must be a directory")
    if os.listdir(folder_name) == 0:
        raise ValueError("ERROR: bad input. Data folder should contain files")
    if not str(workbook_name).endswith('.xlsx'):
        raise ValueError("ERROR: workbook_name input must end with '.xlsx'")
    
    # Set arrays here so we make sure they are empty and we can add to them later
    jaina_logs = []
    varian_logs = []
    slit = None

    # Set directory paths based on pre-defined directory structure, get a list of all files in directories so we can look for the ones we need later
    wavenote_directory = folder_name + '/'
    jaina_directory = folder_name + '/../../../.././ARPES Log Data/Jaina Cadillac/'
    varian_directory = folder_name + '/../../../.././ARPES Log Data/Varian Cadillac/'
    wavenote_names = os.listdir(wavenote_directory)
    jaina_names = os.listdir(jaina_directory)
    varian_names = os.listdir(varian_directory)

    # Organize .pxt files to be in the correct order so when values are printed they are in numerical order
    for name in wavenote_names:
        if not '.pxt' in name:
            wavenote_names.remove(name)
    sorted_waves = sorted(
        wavenote_names,
        key=lambda x: float(re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", x)[-1]) if re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", x) else 0
        )   
    
    # Make sure files in directories are just the files we want so incompatible files are not read
    for f in jaina_names:
        if f.endswith('.log'):
            jaina_logs = jaina_logs + [jaina_directory + f]
    for f in varian_names:
        if f.endswith('.log'):
            varian_logs = varian_logs + [varian_directory + f]

    # Build the workbook with the premade format to hold the data
    build_arpes_workbook(workbook_name)

    # Call insert_scan_rows on each .pxt file in our ordered list to get all the data from the file and insert it into our open workbook
    for f in sorted_waves:
        if f.endswith('.pxt'):
            new_slit = insert_scan_row(wavenote_directory + f,jaina_logs[0],varian_logs[0],workbook_name)
            if slit == None:
                slit = new_slit
            elif slit != new_slit and abs(float(slit) - float(new_slit)) > 7:
                analyzer_slit_row(new_slit, workbook_name)
                insert_scan_row(wavenote_directory + f,jaina_logs[0],varian_logs[0],workbook_name)
                slit = new_slit
    

def single_log_grapher(log_file, scan_folder, log_type, value):
    """
    ``single_log_grapher`` is a function that take a single Varian or Jaina log file from an ARPES scan folder and plots it. The x axis of the plot is 24 hours, the length of time of the log file, and the y axis is the value that is selected in the ``value`` variable. All scans over the time of the log will be plotted. Each scan will be in a different color. This function has been designed to handle 'Jaina' and 'Varian' log files, the types existing on the system for which this code was written. These logs may not exist, or may not be the same for all systems. 

    :args: ``log_file`` is a string or a path to the log file that will be graphed, ending in '.log'. ``scan_folder`` is a string or a path to the folder that contains the scans that happen during the log time.``log_type`` is a string that is either 'varian' or 'jaina'.``value`` is a string that is the log value that will be graphed.
    
    :return: Does not return anything. Plots, displays, and saves a .png file of the plot.
    
    :exceptions: ``log_file`` must be a .log file. ``scan_folder`` must be a folder.  ``scan_folder`` must have files in it. ``log_type`` must be either 'varian' or 'jaina'. ``value`` must be a valid value. 
    """

    # Make sure all inputs are correct types and have data to be read so function doesn't break
    if not log_file.endswith('.log'):
        raise ValueError("ERROR: log file must end with '.log'")
    if os.path.isfile(log_file) is False:
        raise ValueError("ERROR: bad input. Expected file")
    if not os.path.isdir(scan_folder):
        raise ValueError("ERROR: scan folder must be a folder")
    if os.listdir(scan_folder) == 0:
        raise ValueError("ERROR: bad input. Data folder should contain files")
    if log_type.lower() != 'jaina' and log_type.lower() != 'varian':
        raise ValueError("ERROR: Log type must be Jaina or Varian")
    if not (value in jaina_values or value in varian_values):
        raise ValueError("ERROR: Value must be a Jaina or Varian value")

    # Value index set to one so we can use it to figure out the index of our desired value later, log_lines set so it can hold lines later, segment_times set so it can hold data later
    value_index = 0
    log_lines = []  
    segment_times = []

    # Set colors and known possible values to be read from log files so they can be referenced later
    colors = ['red', 'blue', 'orange', 'green', 'purple', 'yellow', 'brown', 'pink', 'light blue', 'beige', 'light green']
    jaina_values = ['Timestamp', 'DiodeA', 'DiodeB', 'Heater', 'HeaterSetPoint', 'HeaterRange', 'OutputMode', 'RampMode', 'RampRate', 
                    'ZoneRampRate', 'CryoTemp', 'CryoLSetPt', 'CryoHSetPt', 'IG_Val', 'ARPES_IG', 'ARPES_PG1', 'ARPES_PG2', 
                    'Mono_IG', 'Mono_PG1', 'Mono_PG2', 'SD_IG', 'SD_PG1', 'SD_PG2', 'TC_IG', 'TC_PG1', 'TC_PG2']
    varian_values = ['Timestamp', 'X_status', 'X', 'Y_status', 'Y', 'Z_status', 'Z', 'Theta_status', 'Theta', 'Phi_status', 
                     'Phi', 'Omega', 'ManipLimitCheck', 'ManipSES', 'ARPES_Slit']
    
    # Either jaina_values or varian_values as our list of values based on which file type were given. If it's neither break because something has gone wrong
    if log_type.lower() == 'jaina':
        values = jaina_values
    elif log_type.lower() == 'varian':
        values = varian_values
    else:
        raise ValueError("Invalid log type")

    # Find value index so we know which value we are graphing and can extract data from the correct column. Break if value not found because something has gone wrong
    for i, item in enumerate(values):
        if item == value:
            value_index = i
            break
    else:
        raise ValueError(f"Value '{value}' not found in log type '{log_type}'")

    # Open the file and extract lines into an array of data so we can extract the column we need
    with open(log_file, 'r') as file:
        log_lines = file.readlines()

    # Extract data from the desired column and the timestamp from the first column. Logs are space separated so we can use split()
    value_data = [line.split()[value_index + 1] for line in log_lines]  
    time_data = [line.split()[1] for line in log_lines[1:]] 

    # Convert time to seconds so it can be plotted. Convert dates to desired date format so they can be compared to .pxt start and end dates.
    new_time = [get_sec(item) for item in time_data if item]
    length = len(log_lines) - 2
    log_start_date = log_lines[1].split()[0][:10]
    log_end_date = log_lines[length].split()[0][:10]
    adjusted_log_start_date = log_start_date[6:10] + '-' + log_start_date[0:2] + '-' + log_start_date[3:5]
    adjusted_log_end_date = log_end_date[6:10] + '-' + log_end_date[0:2] + '-' + log_end_date[3:5]

    # Filter .pxt files to make sure we are only reading through scans and no other files
    waves = [os.path.join(scan_folder, name) for name in os.listdir(scan_folder) if name.endswith('.pxt')]

    # Go through each wave in the wave folder and check if the scan happens within the time of the log file
    for item in waves:
        try:

            # Get the start and end times from the scan and read the data from the scan so we can plot it if the scan is in the log time
            end_time = os.path.getmtime(item)
            scan_end = datetime.fromtimestamp(end_time)
            dataset = htmdec_formats.ARPESDataset.from_file(item)
            lines = dataset._metadata.split("\n")

            # Start date, start time, [End data, end time] from lines enerated above.
            scan_end_str = scan_end.strftime('%Y-%m-%d %H:%M:%S')
            scan_times = [lines[28].split('=')[1].strip(), lines[29].split('=')[1].strip(), scan_end_str.split(' ')]

            # For all scan times, check if the scan is within time bounds. If it is, put it in segment times. If it overlaps either bound, add a bounded start/end time so all data is within the plot x axis
            if scan_times[0] == adjusted_log_start_date and scan_times[2][0] == adjusted_log_end_date:
                segment_times.append((scan_times[1], scan_times[2][1]))
            elif scan_times[0] == adjusted_log_start_date:
                segment_times.append((scan_times[1], '23:59:59'))
            elif scan_times[2][0] == adjusted_log_end_date:
                segment_times.append(('00:00:01', scan_times[2][1]))
                
        except Exception as e:
            print(f"An unexpected error occurred in single log grapher item loop: {e}")
    
    # Convert time and value data into numpy arrays
    x_values = np.array(new_time) / 3600
    y_values = np.array(value_data[1:], dtype=float)  # Ensure y_values are all numeric
    
    # Set end time and start time variables so we can reference them when reading data
    previous_end_time = None
    first_start_time = get_sec(segment_times[0][0]) / 3600

    # Check if the first scan starts after the log start time. If it does plot a line up to the first scan titled 'Initial gap' 
    if x_values[0] < first_start_time:
        gap_mask = (x_values < first_start_time)
        plt.plot(x_values[gap_mask], y_values[gap_mask], linestyle='-', color='grey', label='Initial Gap')
        plt.scatter(x_values[gap_mask], y_values[gap_mask], color='grey', edgecolor='grey')

    # Loop through all scan times recorded in segment_times and plot them
    for i, (start_time, end_time) in enumerate(segment_times):

        # Convert seconds to hours for plotting
        start_sec = get_sec(start_time) / 3600
        end_sec = get_sec(end_time) / 3600
        
        # If there is a gap between scans, plot a grey line there and label it 'Gap'
        if previous_end_time is not None and start_sec > previous_end_time:
            gap_mask = (x_values > previous_end_time) & (x_values < start_sec)
            plt.plot(x_values[gap_mask], y_values[gap_mask], linestyle='-', color='grey', label='Gap')
            plt.scatter(x_values[gap_mask], y_values[gap_mask], color='grey', edgecolor='grey')

        # Set the x_values to the time of the scan, and the y values to the selected measurement values
        segment_mask = (x_values >= start_sec) & (x_values <= end_sec)
        x_segment = x_values[segment_mask]
        y_segment = y_values[segment_mask]

        # Plot x and y values created above as a line segement, label it 'Scan X' and give it a unique color
        plt.plot(x_segment, y_segment, linestyle='-', color=colors[i], label=f'Scan {i+1}')
        plt.scatter(x_segment, y_segment, color=colors[i], edgecolor=colors[i])

        # Set the last end time to the end of the most recent scan
        previous_end_time = end_sec

    # Final segment for values greater than the last end time
    final_mask = x_values >= get_sec(segment_times[-1][1]) / 3600
    x_segment = x_values[final_mask]
    y_segment = y_values[final_mask]

    # Plot legend and save the figure
    plt.legend(bbox_to_anchor=(1.04, 1), loc="upper left")
    plt.xlabel('Time (hours)')
    plt.ylabel('Y Values')
    
    # Plot is saved as 'plot.png', plot is shown so the user can see it
    plt.savefig('plot.png', bbox_inches='tight', dpi=500)
    plt.show()


def arpes_previewer(pxt_file):
    """
    ``previewer`` is a function that takes a pxt file as an input and displays the scan data and metadata from that file.

    :args: ``pxt_file`` is an ARPES .pxt file that contains data from an ARPES scan.

    :return: Does not return anything. Displays data and metadata from the file.

    :exceptions: Will throw an exception if the input``pxt_file`` is not a .pxt file.
    """

    # Make sure the input file is a .pxt file so the function can run correctly
    if not pxt_file.endswith('.pxt'):
        raise ValueError("ERROR: input file must end with '.pxt'")
    if os.path.isfile(pxt_file) is False:
        raise ValueError("ERROR: bad input. Expected file")
    if os.path.getsize(pxt_file) < 10:
        raise ValueError("ERROR: This size of file cannot be handled by this function. File too small.")

    # Initialize QtWidgets so we can make a widget that holds data
    app = QApplication([])
    window = QWidget()
    layout = QVBoxLayout()

    # Run htmdec commands as a subprocess to get data into hdf5 form so we can get all data, there might be a more efficient way to do this
    command = [
        "htmdec-formats", 
        "pxt-to-hdf5", 
        pxt_file, 
        "data_holder.hdf5"
    ]
    # Execute the command
    subprocess.run(command, check=True)

    # From the created .hdf5 file, read out the intensity map data and display it. This is part of the data the we want to display without the user having to open the .pxt file
    filename = "data_holder.hdf5"
    with h5py.File(filename, "r") as f:
        a_group_key = list(f.keys())[0]
        ds_arr = f[a_group_key][()]  # returns as a numpy array
    ds_arr = np.squeeze(ds_arr)
    os.remove(filename)
    plt.imshow(ds_arr, cmap='gray', interpolation='nearest') # Intensity map is just grey scale intensity
    plt.title('Greyscale Intensity Map')
    plt.show()

    # Read out metadata and break it up into lines
    dataset = htmdec_formats.ARPESDataset.from_file(pxt_file)
    lines = dataset._metadata.split("\n")

    # Extract the name of the file so we can display it in the widget 
    display_label = pxt_file.split('/')
    if len(display_label) > 1:
        final_label = display_label[-1]
    else:
        final_label = display_label
    label = QLabel(final_label + ' Metadata')

    # Create a QTextEdit for displaying multi-line text
    text_edit = QTextEdit()
    text_edit.setPlainText(''.join(lines))  # Join list of lines into a single string

    # Add widgets to layout
    layout.addWidget(label)
    layout.addWidget(text_edit)
    window.setLayout(layout)
    
    # Show the window
    window.show()
    app.exec_()

def analyzer_slit_row(slit, workbook_name):
    """
    ``analyzer_slit_row`` is a function that creates a row in an excel .xlsx file with a specific premade format containing the slit information. This row is inserted each time there is a change the analyzer slit value. 

    :args: ``slit`` a string that is the value of the analyzer slit that is passed. ``workbook_name`` is a string or a path that must end with '.xlsx'.

    :return: Does not return anything. Creates a row within an excel file. 

    :exceptions: ``workbook_name`` must end with '.xslx'.
    
    """

    if not str(workbook_name).endswith('.xlsx'):
        raise ValueError("ERROR: workbook_name input must end with '.xlsx'")

    # Initialize known workbook values, set specific color
    starting_row = 6
    starting_cell = 0
    first_col = 1
    light_yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")#light yellow

    wb = wb = load_workbook(filename = workbook_name)
    ws = wb.active

    # Find first empty row
    while starting_cell == 0:
        cell = ws.cell(row=starting_row, column=first_col)
        if not cell.value:
            starting_cell = 1
        else:
            starting_row += 1
    starting_row -= 1

    adjusted_slit = round(float(slit), -2)
    if adjusted_slit == 100:
        slit_text = '100 (0.1n)'
    elif adjusted_slit == 200:
        slit_text = '200 (0.2n)'
    elif adjusted_slit == 300:
        slit_text = '300 (0.3n)'
    elif adjusted_slit == 400:
        slit_text = '400 (0.2a)'
    elif adjusted_slit == 500:
        slit_text = '500 (0.3a)'
    elif adjusted_slit == 600:
        slit_text = '600 (0.5a)'
    elif adjusted_slit == 700:
        slit_text = '700 (0.8a)'
    elif adjusted_slit == 800:
        slit_text = '800 (1.5a)'
    elif adjusted_slit == 900:
        slit_text = '900 (2.5a)'

    # Merge all the cells so we have one big break in the rows, input and format text
    ws.merge_cells(f'A{starting_row}:X{starting_row}')
    cell = ws[f'A{starting_row}']
    cell.value = '* * *  Analyzer Slit Value = ' + slit_text + '  * * *' # Actual text
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    font = Font(size=14)  # Set font size 
    ws[f'A{starting_row}'].font = font
    ws[f'A{starting_row}'].fill = light_yellow_fill

    wb.save(workbook_name)


def main():
    parser = argparse.ArgumentParser(description="Process some input files for background and sample data")
    parser.add_argument("folder_name", help="the input file for the background data")
    parser.add_argument("workbook_name", help="the input file for the sample data")
    args = parser.parse_args()
    arpes_folder_workbook(args.folder_name, args.workbook_name)


if __name__ == "__main__":
    main()